import urllib3
import urllib
import http.cookiejar as cookielib
import re
import time
import numpy as np
import openpyxl
import random


# https://blog.csdn.net/anlun/article/details/43404661
# http://outofmemory.cn/code-snippet/14012/DiscuzRobot-Discuz-machine-program-achieve--and-kind-function


class Retriver:
    def __init__(self, filename, sheetname, startrow=2):
        self.filename = filename
        self.sheetname = sheetname
        self.wb = openpyxl.load_workbook(filename, data_only=True)
        self.row = startrow

    def addPostToList(self, url, subject, content):
        sheet = self.wb[self.sheetname]
        sheet.cell(row=self.row, column=1).value = subject
        print(content)
        sheet.cell(row=self.row, column=2).value = content
        sheet.cell(row=self.row, column=3).value = url
        sheet.cell(row=self.row, column=4).value = 1
        self.wb.save(self.filename)
        self.row += 1

    def retrivePosts(self):
        sheet = self.wb[self.sheetname]
        posts = []
        row = self.row
        title = sheet.cell(row=row, column=1).value
        while title != "" and title is not None:
            content = sheet.cell(row=row, column=2).value
            url = sheet.cell(row=row, column=3).value
            isnew = sheet.cell(row=row, column=4).value
            username = sheet.cell(row=row, column=5).value
            password = sheet.cell(row=row, column=6).value
            fid = sheet.cell(row=row, column=7).value

            article = Article(title.strip(), content, url, isnew,
                              username, password, fid)
            posts.append(article)

            row += 1
            title = sheet.cell(row=row, column=1).value

        return posts


class Article:
    def __init__(self, subject, content, url, isnew, username, password, fid):
        self.subject = subject
        self.content = content
        self.url = url
        self.isnew = isnew
        self.username = username
        self.password = password
        self.fid = fid


class DiscuzRobot:

    def __init__(self, forumUrl, userName, password, proxy=None):
        ''' 初始化论坛url、用户名、密码和代理服务器 '''
        self.forumUrl = forumUrl
        self.userName = userName
        self.password = password
        self.formhash = ''
        self.isLogon = False
        self.isSign = False
        self.xq = ''

        self.jar = cookielib.CookieJar()
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.jar))

        self.hdr = {
            'Accept': '*/*',
            'Origin': 'https://www.kesucorp.com',
            'X-Requested-With': 'XMLHttpRequest',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36',
            'Content-Type': 'application/json; charset=UTF-8',
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'Referer': self.forumUrl,
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.8',

            # 'Cookie': self.cookie
        }

        urllib.request.install_opener(self.opener)

    def login(self):
        ''' 登录论坛 '''
        url = self.forumUrl + \
            "/member.php?mod=logging&action=login&loginsubmit=yes&infloat=yes&lssubmit=yes&inajax=1"
        postData = urllib.parse.urlencode({'username': self.userName, 'password': self.password,
                                           'handlekey': 'ls',  'quickforward': 'yes',  'fastloginfield': 'username'})
        req = urllib.request.Request(url, data=postData.encode(
            encoding='UTF8'), method='POST', headers=self.hdr)

        content = self.opener.open(req)
        # print(self.jar)
        self.opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.jar))

        content = self.opener.open(req).read().decode('UTF8')

        # print(content)
        if self.userName in content:
            self.isLogon = True
            print('logon success!')
            self.initFormhashXq()
        else:
            print('logon faild!')

    def initFormhashXq(self):
        ''' 获取formhash和心情 '''
        content = urllib.request.urlopen(
            self.forumUrl + '/plugin.php?id=dsu_paulsign:sign').read().decode('UTF8')
        rows = re.findall(
            r'input type=\"hidden\" name=\"formhash\" value=\"(.*?)\"', content)
        if len(rows) != 0:
            self.formhash = rows[0]
            print('formhash is: ' + self.formhash)
        else:
            print('none formhash!')
        rows = re.findall(
            r'input id=.* type=\"radio\" name=\"qdxq\" value=\"(.*?)\"', content)

    def reply(self, fid, tid, subject=u'', msg=u'支持~~~顶一下下~~嘻嘻'):
        ''' 回帖 '''
        url = self.forumUrl + \
            '/forum.php?mod=post&action=reply&fid={}&tid={}&extra=page%3D1&replysubmit=yes&infloat=yes&handlekey=fastpost&inajax=1'.format(
                fid, tid)
        postData = urllib.parse.urlencode({'formhash': self.formhash, 'message': msg.encode(
            'UTF8'), 'subject': subject, 'posttime': int(time.time()), 'usesig': '1'})
        req = urllib.request.Request(url, data=postData.encode(
            encoding='UTF8'), method='POST')
        content = urllib.request.urlopen(req).read().decode('UTF8')
        # print content
        if u'发布成功' in content:
            print('reply success!')
            return True
        else:
            print(content)
            print('reply faild!')
            return False

    def publish(self, fid, subject=u'asasdfasdfasdfdfasdf', msg=u'asdfasdfasdfasdf'):
        ''' 发帖 '''
        url = self.forumUrl + \
            '/forum.php?mod=post&action=newthread&fid={}&extra=&topicsubmit=yes'.format(
                fid)
        print(self.formhash)
        postData = urllib.parse.urlencode({'formhash': self.formhash, 'message': msg, 'subject': subject,
                                           'posttime': int(time.time())-100, 'addfeed': '1', 'allownoticeauthor': '1',
                                           'save': '',  'usesig': '1', 'wysiwyg': '0'})

        req = urllib.request.Request(url, data=postData.encode(
            encoding='UTF8'), method='POST', headers=self.hdr)
        # content = urllib.request.urlopen(req).read().decode('UTF8')
        content = self.opener.open(req).read().decode('UTF8')
        print(subject)
        if subject in content:
            print('publish success!')
            rows = re.findall(r'mod=viewthread&tid=(.*?)\"', content)
            return True, rows[0]
        else:
            print('publish faild!')
            print(content)
            return False, -1


if __name__ == '__main__':

    retriver = Retriver("contentxls\content.xlsx", "Sheet1", 2)
    password = 'kingva584520'
    lsittopublish = retriver.retrivePosts()
    siteurl = 'http://d2.zhaiquanren.net'
    tid = -1
    faildlist = []
    maxtrytosubmit = 20

    for aticle in lsittopublish[0:36]:
        print(aticle.username)
        robot = DiscuzRobot(siteurl,
                            aticle.username, password)
        needtosubmit = True

        publishtimes = 0
        print(aticle.isnew)

        time.sleep(random.randint(6, 108))

        if aticle.isnew == 1:
            tid = -1

            hrefurl = "\r\n\r\n[url=" + \
                aticle.url+"][color=#ff0000][b]点此查看文章详情" + \
                ''+"[/b][/color][/url]" if aticle.url else ""

            content = aticle.content + hrefurl

            while needtosubmit and publishtimes <= maxtrytosubmit:
                robot.login()
                if robot.isLogon:
                    submitted, temptid = robot.publish(
                        aticle.fid, aticle.subject, content)
                    needtosubmit = not submitted
                    if submitted:
                        tid = temptid
                    else:
                        if publishtimes == maxtrytosubmit:
                            faildlist.append(aticle.subject)
                        publishtimes += 1

                    print(needtosubmit)
                    print(tid)

        elif aticle.isnew == 0:

            while needtosubmit and publishtimes <= maxtrytosubmit:
                robot.login()
                if robot.isLogon:
                    submitted = robot.reply(
                        aticle.fid, tid, " ", aticle.content)
                    needtosubmit = not submitted

                    if not submitted:
                        if publishtimes == maxtrytosubmit:
                            faildlist.append(aticle.subject)
                        publishtimes += 1

                    print(needtosubmit)
                    print(tid)

    for faileditem in faildlist:
        print(faileditem)

    # robot.reply(107137)
